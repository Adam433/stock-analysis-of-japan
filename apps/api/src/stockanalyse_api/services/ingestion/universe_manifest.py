from __future__ import annotations

import csv
import re
from pathlib import Path
from urllib.parse import urljoin
from urllib.request import Request
from urllib.request import urlopen

JPX_LISTED_ISSUES_PAGE_URL = "https://www.jpx.co.jp/markets/statistics-equities/misc/01.html"
DEFAULT_USER_AGENT = "stockAnalyse/0.1 universe-sync"

MARKET_HEADER_CANDIDATES = ("市場・商品区分", "市場区分", "market", "market/product")
CODE_HEADER_CANDIDATES = ("コード", "銘柄コード", "code")

TSE_COMMON_STOCK_INCLUDE_KEYWORDS = ("プライム", "スタンダード", "グロース", "prime", "standard", "growth")
TSE_COMMON_STOCK_EXCLUDE_KEYWORDS = (
    "etf",
    "etn",
    "reit",
    "インフラファンド",
    "ベンチャーファンド",
    "tokyo pro market",
    "pro market",
)


def fetch_latest_jpx_listed_issues_workbook_url(page_url: str = JPX_LISTED_ISSUES_PAGE_URL) -> str:
    request = Request(page_url, headers={"User-Agent": DEFAULT_USER_AGENT})
    with urlopen(request) as response:
        html = response.read().decode("utf-8", errors="ignore")

    match = re.search(r'href="([^"]+\.(?:xls|xlsx|csv))"', html, flags=re.IGNORECASE)
    if not match:
        raise ValueError("Could not locate a listed-issues workbook link on the JPX page.")

    return urljoin(page_url, match.group(1))


def download_file(url: str, destination: Path) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    request = Request(url, headers={"User-Agent": DEFAULT_USER_AGENT})
    with urlopen(request) as response:
        destination.write_bytes(response.read())
    return destination


def _normalize_header(value: object) -> str:
    return str(value).strip().lower()


def _pick_column(headers: list[str], candidates: tuple[str, ...]) -> str:
    normalized = {_normalize_header(header): header for header in headers}
    for candidate in candidates:
        if candidate.lower() in normalized:
            return normalized[candidate.lower()]
    raise ValueError(f"Required header not found. Tried: {', '.join(candidates)}")


def load_rows_from_table(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return list(csv.DictReader(handle))

    if suffix == ".xlsx":
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value).strip() for value in rows[0]]
        return [
            {headers[index]: "" if value is None else str(value) for index, value in enumerate(row)}
            for row in rows[1:]
        ]

    if suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
        records: list[dict[str, str]] = []
        for row_index in range(1, sheet.nrows):
            row = {
                headers[col_index]: str(sheet.cell_value(row_index, col_index)).strip()
                for col_index in range(sheet.ncols)
            }
            records.append(row)
        return records

    raise ValueError(f"Unsupported workbook type: {path.suffix}")


def _is_tse_common_stock(market_label: str) -> bool:
    normalized = market_label.strip().lower()
    if not normalized:
        return False
    if any(keyword in normalized for keyword in TSE_COMMON_STOCK_EXCLUDE_KEYWORDS):
        return False
    return any(keyword in normalized for keyword in TSE_COMMON_STOCK_INCLUDE_KEYWORDS)


def normalize_symbol(code: str) -> str | None:
    raw = code.strip().upper()
    if not raw:
        return None
    if raw.endswith(".0") and re.fullmatch(r"\d+\.0", raw):
        raw = raw[:-2]
    if raw.endswith(".T"):
        return raw
    if re.fullmatch(r"[0-9A-Z]+", raw) is None:
        return None
    return f"{raw}.T"


def build_tse_common_stock_symbols(rows: list[dict[str, str]]) -> list[str]:
    if not rows:
        return []

    headers = list(rows[0].keys())
    market_header = _pick_column(headers, MARKET_HEADER_CANDIDATES)
    code_header = _pick_column(headers, CODE_HEADER_CANDIDATES)

    symbols: set[str] = set()
    for row in rows:
        if not _is_tse_common_stock(row.get(market_header, "")):
            continue
        symbol = normalize_symbol(row.get(code_header, ""))
        if symbol:
            symbols.add(symbol)

    return sorted(symbols)


def write_symbol_manifest(symbols: list[str], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("".join(f"{symbol}\n" for symbol in symbols), encoding="utf-8")
    return output_path
